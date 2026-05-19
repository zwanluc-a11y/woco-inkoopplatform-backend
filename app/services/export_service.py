from __future__ import annotations

"""
Export Service

Generates formatted Excel exports for:
1. Spend analysis (pivot table)
2. Risk assessment report
3. Procurement calendar
"""

import io
import logging
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.models.contract import Contract, ContractSupplier
from app.models.procurement_calendar_item import ProcurementCalendarItem
from app.models.risk_assessment import RiskAssessment
from app.models.supplier import Supplier
from app.models.supplier_categorization import SupplierCategorization
from app.models.supplier_yearly_spend import SupplierYearlySpend

logger = logging.getLogger(__name__)

# Styling constants
NAVY = "1D2F45"
GOLD = "C8A45A"
RED = "E8313A"
GREEN = "10B981"
YELLOW = "F59E0B"
ORANGE = "F97316"
WHITE = "FFFFFF"
LIGHT_GRAY = "F3F4F6"

HEADER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color=WHITE, size=11)
GOLD_FILL = PatternFill(start_color=GOLD, end_color=GOLD, fill_type="solid")
GOLD_FONT = Font(name="Calibri", bold=True, color=NAVY, size=11)
GREEN_FILL = PatternFill(start_color=GREEN, end_color=GREEN, fill_type="solid")
YELLOW_FILL = PatternFill(start_color=YELLOW, end_color=YELLOW, fill_type="solid")
RED_FILL = PatternFill(start_color=RED, end_color=RED, fill_type="solid")
ORANGE_FILL = PatternFill(start_color=ORANGE, end_color=ORANGE, fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)
CURRENCY_FORMAT = '#,##0'
PERCENT_FORMAT = '0%'


class ExportService:
    def __init__(self, db: Session):
        self.db = db

    def _get_category_label(self, org_id: int) -> str:
        """Header label for the category column."""
        return "Inkooppakket"

    def export_spend_analysis(self, org_id: int) -> io.BytesIO:
        """Export spend analysis as formatted Excel with 3 sheets: Leveranciers, Categorieën, Inzichten."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Leveranciers"

        # Get all years
        years_q = (
            self.db.query(SupplierYearlySpend.year)
            .filter(SupplierYearlySpend.organization_id == org_id)
            .distinct()
            .order_by(SupplierYearlySpend.year)
            .all()
        )
        years = [r.year for r in years_q]

        # Build supplier→contract name mapping
        contract_supplier_rows = (
            self.db.query(ContractSupplier.supplier_id, Contract.name)
            .join(Contract, Contract.id == ContractSupplier.contract_id)
            .filter(Contract.organization_id == org_id)
            .all()
        )
        supplier_contracts: dict[int, list[str]] = {}
        for row in contract_supplier_rows:
            supplier_contracts.setdefault(row.supplier_id, []).append(row.name)
        has_any_contracts = len(supplier_contracts) > 0

        # Headers
        cat_label = self._get_category_label(org_id)
        headers = ["#", "Leverancier", "Leverancierscode", cat_label]
        if has_any_contracts:
            headers.append("Contract")
        headers.extend([str(y) for y in years])
        headers.append("Totaal")

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER

        # Data
        suppliers = (
            self.db.query(Supplier)
            .filter(Supplier.organization_id == org_id)
            .all()
        )

        # Build spend data
        rows = []
        for s in suppliers:
            spend_by_year = {ys.year: float(ys.total_amount) for ys in s.yearly_spends}
            total = sum(spend_by_year.values())
            cat_name = ""
            cats = s.categorizations or []
            if cats:
                cat_parts = []
                for c in cats:
                    if c.category:
                        if len(cats) > 1:
                            cat_parts.append(f"{c.category.inkooppakket} ({c.percentage:.0f}%)")
                        else:
                            cat_parts.append(c.category.inkooppakket)
                cat_name = " | ".join(cat_parts)
            contract_names = supplier_contracts.get(s.id, [])
            rows.append({
                "name": s.name,
                "code": s.supplier_code or "",
                "category": cat_name,
                "contract": " | ".join(contract_names) if contract_names else "",
                "has_contract": len(contract_names) > 0,
                "spends": spend_by_year,
                "total": total,
            })

        # Sort by total descending (absolute)
        rows.sort(key=lambda x: abs(x["total"]), reverse=True)

        # Column offset for year columns depends on whether contract column exists
        year_col_start = 6 if has_any_contracts else 5

        for row_idx, row in enumerate(rows, 2):
            ws.cell(row=row_idx, column=1, value=row_idx - 1).border = THIN_BORDER
            ws.cell(row=row_idx, column=2, value=row["name"]).border = THIN_BORDER
            ws.cell(row=row_idx, column=3, value=row["code"]).border = THIN_BORDER
            ws.cell(row=row_idx, column=4, value=row["category"]).border = THIN_BORDER

            if has_any_contracts:
                contract_cell = ws.cell(row=row_idx, column=5, value="Ja" if row["has_contract"] else "Nee")
                contract_cell.alignment = Alignment(horizontal="center")
                contract_cell.border = THIN_BORDER
                if row["has_contract"]:
                    contract_cell.fill = GREEN_FILL
                    contract_cell.font = Font(bold=True, color=WHITE)

            for yr_idx, yr in enumerate(years):
                cell = ws.cell(
                    row=row_idx, column=year_col_start + yr_idx,
                    value=row["spends"].get(yr, 0)
                )
                cell.number_format = CURRENCY_FORMAT
                cell.border = THIN_BORDER

            total_cell = ws.cell(
                row=row_idx, column=year_col_start + len(years),
                value=row["total"]
            )
            total_cell.number_format = CURRENCY_FORMAT
            total_cell.font = Font(bold=True)
            total_cell.border = THIN_BORDER

            # Alternate row coloring
            if row_idx % 2 == 0:
                for col in range(1, len(headers) + 1):
                    cell = ws.cell(row=row_idx, column=col)
                    # Don't overwrite green contract fill
                    if not (has_any_contracts and col == 5 and row["has_contract"]):
                        cell.fill = PatternFill(
                            start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid"
                        )

        # Column widths
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 35
        if has_any_contracts:
            ws.column_dimensions["E"].width = 12
        for i in range(len(years) + 1):
            ws.column_dimensions[get_column_letter(year_col_start + i)].width = 15

        # Freeze panes
        freeze_col = "F" if has_any_contracts else "E"
        ws.freeze_panes = f"{freeze_col}2"

        # ── Sheet 2: Categorieën ────────────────────────────────────────
        self._add_categories_sheet(wb, org_id, years)

        # ── Sheet 3: Inzichten ──────────────────────────────────────────
        self._add_insights_sheet(wb, org_id)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    # ── Helper: Categorieën sheet (flat, filterable) ────────────────────
    def _add_categories_sheet(self, wb: Workbook, org_id: int, years: list[int]) -> None:
        """Add a sheet with one row per supplier-in-category combo, fully filterable via AutoFilter.

        Each multi-category supplier appears once per category with their actual
        weighted percentage (e.g. 60% / 40% for a supplier split across two cats).
        Use Excel's AutoFilter on the Categorie column to filter to one category.
        """
        from app.services.insights_service import InsightsService

        ws = wb.create_sheet("Categorieën")
        svc = InsightsService(self.db)
        try:
            pivot = svc.get_category_pivot(org_id)
        except Exception as e:
            logger.warning("Could not load category pivot for export: %s", e)
            pivot = {"categories": []}

        categories = pivot.get("categories", [])

        # Headers — flat columns so AutoFilter works
        headers = ["Categorie", "Leverancier", "Aandeel %"]
        headers.extend([str(y) for y in years])
        headers.append("Totaal")

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER

        # Build flat rows: one per (category, supplier) combination
        flat_rows: list[dict] = []
        for cat in categories:
            cat_name = cat.get("category_name") or "Ongecategoriseerd"
            for s in cat.get("suppliers", []):
                flat_rows.append({
                    "category": cat_name,
                    "name": s["name"],
                    "percentage": s.get("percentage"),  # None for ongecategoriseerd
                    "spends": s.get("spends", {}),
                    "total": s.get("total", 0),
                })

        # Sort by category name, then by spend descending within category
        flat_rows.sort(key=lambda r: (r["category"], -abs(r["total"])))

        for row_idx, r in enumerate(flat_rows, 2):
            ws.cell(row=row_idx, column=1, value=r["category"]).border = THIN_BORDER
            ws.cell(row=row_idx, column=2, value=r["name"]).border = THIN_BORDER
            pct = r["percentage"]
            # Store as a number with % format so Excel filters/sorts numerically
            pct_cell = ws.cell(row=row_idx, column=3,
                               value=(float(pct) / 100.0) if pct is not None else None)
            pct_cell.number_format = "0%"
            pct_cell.alignment = Alignment(horizontal="center")
            pct_cell.border = THIN_BORDER
            for yr_idx, yr in enumerate(years):
                cell = ws.cell(row=row_idx, column=4 + yr_idx,
                               value=r["spends"].get(str(yr), 0))
                cell.number_format = CURRENCY_FORMAT
                cell.border = THIN_BORDER
            total_cell = ws.cell(row=row_idx, column=4 + len(years), value=r["total"])
            total_cell.number_format = CURRENCY_FORMAT
            total_cell.font = Font(bold=True)
            total_cell.border = THIN_BORDER

            # Alternating row colors for readability
            if row_idx % 2 == 0:
                for col in range(1, len(headers) + 1):
                    if col == 3:  # don't overwrite the % alignment
                        ws.cell(row=row_idx, column=col).fill = PatternFill(
                            start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
                    else:
                        ws.cell(row=row_idx, column=col).fill = PatternFill(
                            start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")

        # Enable AutoFilter on header row — user can click the filter icon on
        # the Categorie column to pick one or more categories
        last_col_letter = get_column_letter(len(headers))
        last_row = 1 + len(flat_rows)
        if last_row > 1:
            ws.auto_filter.ref = f"A1:{last_col_letter}{last_row}"

        # Column widths
        ws.column_dimensions["A"].width = 35  # Categorie
        ws.column_dimensions["B"].width = 38  # Leverancier
        ws.column_dimensions["C"].width = 12  # Aandeel %
        for i in range(len(years) + 1):
            ws.column_dimensions[get_column_letter(4 + i)].width = 15

        # Freeze the header row + first 3 columns so they stay visible while scrolling
        ws.freeze_panes = "D2"

        # ── Optional: small summary section above table ──
        # We'll insert it as a second worksheet area is overkill; instead add
        # a totals-per-category sheet for quick overview.
        self._add_category_totals_sheet(wb, categories, years)

    def _add_category_totals_sheet(self, wb: Workbook, categories: list[dict], years: list[int]) -> None:
        """Add a 'Categorie totalen' sheet: one row per category, no nested suppliers."""
        ws = wb.create_sheet("Categorie totalen")

        headers = ["#", "Categorie", "Leveranciers"]
        headers.extend([str(y) for y in years])
        headers.append("Totaal")
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER

        # Sort categories by total descending
        cats_sorted = sorted(categories, key=lambda c: abs(c.get("total", 0)), reverse=True)
        for row_idx, cat in enumerate(cats_sorted, 2):
            ws.cell(row=row_idx, column=1, value=row_idx - 1).border = THIN_BORDER
            ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center")
            ws.cell(row=row_idx, column=2,
                    value=cat.get("category_name") or "Ongecategoriseerd").border = THIN_BORDER
            ws.cell(row=row_idx, column=3, value=cat.get("supplier_count", 0)).border = THIN_BORDER
            ws.cell(row=row_idx, column=3).alignment = Alignment(horizontal="center")
            cat_spends = cat.get("spends", {})
            for yr_idx, yr in enumerate(years):
                cell = ws.cell(row=row_idx, column=4 + yr_idx,
                               value=cat_spends.get(str(yr), 0))
                cell.number_format = CURRENCY_FORMAT
                cell.border = THIN_BORDER
            total_cell = ws.cell(row=row_idx, column=4 + len(years), value=cat.get("total", 0))
            total_cell.number_format = CURRENCY_FORMAT
            total_cell.font = Font(bold=True)
            total_cell.border = THIN_BORDER
            if row_idx % 2 == 0:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col).fill = PatternFill(
                        start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")

        last_col_letter = get_column_letter(len(headers))
        last_row = 1 + len(cats_sorted)
        if last_row > 1:
            ws.auto_filter.ref = f"A1:{last_col_letter}{last_row}"

        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 35
        ws.column_dimensions["C"].width = 14
        for i in range(len(years) + 1):
            ws.column_dimensions[get_column_letter(4 + i)].width = 15
        ws.freeze_panes = "D2"

    # ── Helper: Inzichten sheet ─────────────────────────────────────────
    def _add_insights_sheet(self, wb: Workbook, org_id: int) -> None:
        """Add a sheet with transaction insights: per-year summaries, top-5 lists, per-supplier detail."""
        from app.services.insights_service import InsightsService

        ws = wb.create_sheet("Inzichten")
        svc = InsightsService(self.db)
        try:
            ins = svc.get_spend_insights(org_id)
        except Exception as e:
            logger.warning("Could not load spend insights for export: %s", e)
            ins = {"year_summaries": [], "totals": {}, "top_avg_invoice": [], "top_transactions": [], "suppliers_all": []}

        row_idx = 1

        def write_section_header(title: str):
            nonlocal row_idx
            cell = ws.cell(row=row_idx, column=1, value=title)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            ws.row_dimensions[row_idx].height = 22
            row_idx += 1

        def write_column_headers(headers: list[str]):
            nonlocal row_idx
            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=h)
                cell.fill = GOLD_FILL
                cell.font = GOLD_FONT
                cell.alignment = Alignment(horizontal="center")
                cell.border = THIN_BORDER
            row_idx += 1

        # ── Totals ──
        totals = ins.get("totals", {})
        write_section_header("Totalen over alle jaren")
        write_column_headers(["Metric", "Waarde"])
        totals_rows = [
            ("Totale spend", totals.get("total_spend", 0), CURRENCY_FORMAT),
            ("Aantal transacties", totals.get("total_transactions", 0), "#,##0"),
            ("Aantal leveranciers", totals.get("supplier_count", 0), "#,##0"),
            ("Gemiddelde factuurwaarde", totals.get("avg_invoice", 0), CURRENCY_FORMAT),
        ]
        for label, value, fmt in totals_rows:
            ws.cell(row=row_idx, column=1, value=label).border = THIN_BORDER
            val_cell = ws.cell(row=row_idx, column=2, value=value)
            val_cell.number_format = fmt
            val_cell.font = Font(bold=True)
            val_cell.border = THIN_BORDER
            row_idx += 1
        row_idx += 1  # blank line

        # ── Per jaar ──
        write_section_header("Per jaar")
        write_column_headers(["Jaar", "Spend", "Transacties", "Leveranciers", "Gem. factuurwaarde"])
        for ys in ins.get("year_summaries", []):
            ws.cell(row=row_idx, column=1, value=ys["year"]).border = THIN_BORDER
            ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center")
            cell = ws.cell(row=row_idx, column=2, value=ys["total_spend"])
            cell.number_format = CURRENCY_FORMAT
            cell.border = THIN_BORDER
            cell = ws.cell(row=row_idx, column=3, value=ys["transaction_count"])
            cell.number_format = "#,##0"
            cell.border = THIN_BORDER
            ws.cell(row=row_idx, column=4, value=ys["supplier_count"]).border = THIN_BORDER
            ws.cell(row=row_idx, column=4).alignment = Alignment(horizontal="center")
            cell = ws.cell(row=row_idx, column=5, value=ys["avg_invoice"])
            cell.number_format = CURRENCY_FORMAT
            cell.border = THIN_BORDER
            row_idx += 1
        row_idx += 1

        # ── Top 5 hoogste gem. factuurwaarde ──
        write_section_header("Top 5 hoogste gemiddelde factuurwaarde (min. 3 transacties)")
        write_column_headers(["#", "Leverancier", "Categorie", "Transacties", "Gem. factuur"])
        for i, s in enumerate(ins.get("top_avg_invoice", []), 1):
            ws.cell(row=row_idx, column=1, value=i).border = THIN_BORDER
            ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center")
            ws.cell(row=row_idx, column=2, value=s["name"]).border = THIN_BORDER
            ws.cell(row=row_idx, column=3, value=s.get("category_name") or "—").border = THIN_BORDER
            cell = ws.cell(row=row_idx, column=4, value=s["total_transactions"])
            cell.number_format = "#,##0"
            cell.border = THIN_BORDER
            cell = ws.cell(row=row_idx, column=5, value=s["avg_invoice"])
            cell.number_format = CURRENCY_FORMAT
            cell.font = Font(bold=True)
            cell.border = THIN_BORDER
            row_idx += 1
        row_idx += 1

        # ── Top 5 meeste transacties ──
        write_section_header("Top 5 meeste transacties")
        write_column_headers(["#", "Leverancier", "Categorie", "Transacties", "Spend"])
        for i, s in enumerate(ins.get("top_transactions", []), 1):
            ws.cell(row=row_idx, column=1, value=i).border = THIN_BORDER
            ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center")
            ws.cell(row=row_idx, column=2, value=s["name"]).border = THIN_BORDER
            ws.cell(row=row_idx, column=3, value=s.get("category_name") or "—").border = THIN_BORDER
            cell = ws.cell(row=row_idx, column=4, value=s["total_transactions"])
            cell.number_format = "#,##0"
            cell.font = Font(bold=True)
            cell.border = THIN_BORDER
            cell = ws.cell(row=row_idx, column=5, value=s["total_spend"])
            cell.number_format = CURRENCY_FORMAT
            cell.border = THIN_BORDER
            row_idx += 1
        row_idx += 1

        # ── Detail per leverancier ──
        write_section_header("Detail per leverancier")
        all_suppliers = ins.get("suppliers_all", [])
        years = sorted({int(yr) for s in all_suppliers for yr in s.get("yearly", {}).keys()}) if all_suppliers else []

        detail_headers = ["#", "Leverancier", "Categorie"]
        for yr in years:
            detail_headers.extend([f"{yr} trans.", f"{yr} gem. factuur"])
        detail_headers.extend(["Totaal transacties", "Totaal spend", "Gem. factuur"])
        write_column_headers(detail_headers)

        # Sort by total_spend descending
        suppliers_sorted = sorted(all_suppliers, key=lambda x: abs(x.get("total_spend", 0)), reverse=True)
        for i, s in enumerate(suppliers_sorted, 1):
            ws.cell(row=row_idx, column=1, value=i).border = THIN_BORDER
            ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center")
            ws.cell(row=row_idx, column=2, value=s["name"]).border = THIN_BORDER
            ws.cell(row=row_idx, column=3, value=s.get("category_name") or "—").border = THIN_BORDER
            col = 4
            yearly = s.get("yearly", {})
            for yr in years:
                y = yearly.get(str(yr), {})
                cell = ws.cell(row=row_idx, column=col, value=y.get("transaction_count", 0))
                cell.number_format = "#,##0"
                cell.border = THIN_BORDER
                col += 1
                cell = ws.cell(row=row_idx, column=col, value=y.get("avg_invoice", 0))
                cell.number_format = CURRENCY_FORMAT
                cell.border = THIN_BORDER
                col += 1
            cell = ws.cell(row=row_idx, column=col, value=s.get("total_transactions", 0))
            cell.number_format = "#,##0"
            cell.font = Font(bold=True)
            cell.border = THIN_BORDER
            col += 1
            cell = ws.cell(row=row_idx, column=col, value=s.get("total_spend", 0))
            cell.number_format = CURRENCY_FORMAT
            cell.font = Font(bold=True)
            cell.border = THIN_BORDER
            col += 1
            cell = ws.cell(row=row_idx, column=col, value=s.get("avg_invoice", 0))
            cell.number_format = CURRENCY_FORMAT
            cell.border = THIN_BORDER
            row_idx += 1

        # Column widths
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 38
        ws.column_dimensions["C"].width = 30
        for i in range(4, 4 + len(years) * 2 + 3):
            ws.column_dimensions[get_column_letter(i)].width = 15

    def export_risk_assessment(self, org_id: int, assessment_year: int) -> io.BytesIO:
        """Export risk assessment as formatted Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Risicoanalyse"

        cat_label = self._get_category_label(org_id)
        headers = [
            "#", cat_label, "Groep", "Soort Inkoop",
            "Jaarlijkse Spend", "Aantal Leveranciers",
            "Looptijd (jr)", "Geraamde Opdrachtwaarde",
            "Drempel", "% van Drempel", "Risico",
            "Contract", "Opmerkingen"
        ]

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = THIN_BORDER

        assessments = (
            self.db.query(RiskAssessment)
            .filter(
                RiskAssessment.organization_id == org_id,
                RiskAssessment.assessment_year == assessment_year,
            )
            .order_by(RiskAssessment.estimated_contract_value.desc())
            .all()
        )

        risk_fills = {
            "offertetraject": RED_FILL,
            "meervoudig_onderhands": YELLOW_FILL,
            "enkelvoudig_onderhands": ORANGE_FILL,
            "vrije_inkoop": GREEN_FILL,
        }

        for row_idx, ra in enumerate(assessments, 2):
            cat_name = ra.category.inkooppakket if ra.category else ""
            cat_groep = ra.category.groep if ra.category else ""
            soort = ra.threshold_type or ""
            pct = ra.estimated_contract_value / ra.internal_threshold if ra.internal_threshold else 0

            ws.cell(row=row_idx, column=1, value=row_idx - 1).border = THIN_BORDER
            ws.cell(row=row_idx, column=2, value=cat_name).border = THIN_BORDER
            ws.cell(row=row_idx, column=3, value=cat_groep).border = THIN_BORDER
            ws.cell(row=row_idx, column=4, value=soort).border = THIN_BORDER

            spend_cell = ws.cell(row=row_idx, column=5, value=ra.yearly_spend)
            spend_cell.number_format = CURRENCY_FORMAT
            spend_cell.border = THIN_BORDER

            ws.cell(row=row_idx, column=6, value=ra.supplier_count).border = THIN_BORDER
            ws.cell(row=row_idx, column=7, value=ra.duration_years).border = THIN_BORDER

            ecv_cell = ws.cell(row=row_idx, column=8, value=ra.estimated_contract_value)
            ecv_cell.number_format = CURRENCY_FORMAT
            ecv_cell.font = Font(bold=True)
            ecv_cell.border = THIN_BORDER

            thresh_cell = ws.cell(row=row_idx, column=9, value=ra.internal_threshold)
            thresh_cell.number_format = CURRENCY_FORMAT
            thresh_cell.border = THIN_BORDER

            pct_cell = ws.cell(row=row_idx, column=10, value=pct)
            pct_cell.number_format = PERCENT_FORMAT
            pct_cell.border = THIN_BORDER

            risk_cell = ws.cell(row=row_idx, column=11, value=ra.risk_level)
            risk_cell.fill = risk_fills.get(ra.risk_level, PatternFill())
            risk_cell.font = Font(bold=True, color=WHITE if ra.risk_level != "meervoudig_onderhands" else NAVY)
            risk_cell.alignment = Alignment(horizontal="center")
            risk_cell.border = THIN_BORDER

            ws.cell(row=row_idx, column=12, value="Ja" if ra.has_contract else "Nee").border = THIN_BORDER
            ws.cell(row=row_idx, column=13, value=ra.notes or "").border = THIN_BORDER

        # Column widths
        widths = [6, 40, 25, 15, 15, 12, 10, 20, 15, 12, 15, 10, 30]
        for i, w in enumerate(widths):
            ws.column_dimensions[get_column_letter(i + 1)].width = w

        ws.freeze_panes = "E2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(assessments) + 1}"

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def export_calendar(self, org_id: int) -> io.BytesIO:
        """Export procurement calendar as formatted Excel."""
        wb = Workbook()
        ws = wb.active
        ws.title = "WoCo Inkoopplatform"

        cat_label = self._get_category_label(org_id)
        headers = [
            "#", "Titel", "Prioriteit", cat_label,
            "Geraamde Waarde", "Start voorbereiding",
            "Publicatiedatum", "Status", "Omschrijving"
        ]

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER

        items = (
            self.db.query(ProcurementCalendarItem)
            .filter(ProcurementCalendarItem.organization_id == org_id)
            .order_by(
                ProcurementCalendarItem.priority.desc(),
                ProcurementCalendarItem.target_start_date.asc(),
            )
            .all()
        )

        priority_fills = {
            "high": RED_FILL,
            "medium": YELLOW_FILL,
            "low": GREEN_FILL,
        }

        for row_idx, item in enumerate(items, 2):
            cat_name = item.category.inkooppakket if item.category else ""

            ws.cell(row=row_idx, column=1, value=row_idx - 1).border = THIN_BORDER
            ws.cell(row=row_idx, column=2, value=item.title).border = THIN_BORDER

            prio_cell = ws.cell(row=row_idx, column=3, value=item.priority)
            prio_cell.fill = priority_fills.get(item.priority, PatternFill())
            prio_cell.font = Font(bold=True, color=WHITE if item.priority != "medium" else NAVY)
            prio_cell.alignment = Alignment(horizontal="center")
            prio_cell.border = THIN_BORDER

            ws.cell(row=row_idx, column=4, value=cat_name).border = THIN_BORDER

            val_cell = ws.cell(
                row=row_idx, column=5,
                value=float(item.estimated_value) if item.estimated_value else None
            )
            val_cell.number_format = CURRENCY_FORMAT
            val_cell.border = THIN_BORDER

            ws.cell(
                row=row_idx, column=6,
                value=item.target_start_date
            ).border = THIN_BORDER

            ws.cell(
                row=row_idx, column=7,
                value=item.target_publish_date
            ).border = THIN_BORDER

            ws.cell(row=row_idx, column=8, value=item.status).border = THIN_BORDER
            ws.cell(row=row_idx, column=9, value=item.description or "").border = THIN_BORDER

        # Column widths
        widths = [6, 45, 12, 35, 18, 18, 18, 12, 40]
        for i, w in enumerate(widths):
            ws.column_dimensions[get_column_letter(i + 1)].width = w

        ws.freeze_panes = "C2"

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
